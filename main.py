import pyttsx3
import num2words
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#wb = Workbook()
wb = load_workbook('data pasien.xlsx')
ws = wb.active

input = int(input('Masukkan nama pasien :'))
name = input
temperature = ws["B2"].value
heartrate = ws["A2"].value
systole = ws["C2"].value 
diastole = ws["D2"].value
# print(f'{name}: {temperature}: {heartrate}: {systole}/ {diastole}')

SuhuBadan = temperature
text = num2words.num2words(SuhuBadan, lang='id')
print (f"Suhu tubuh anda adalah {text} derajat celcius")


if SuhuBadan < 36 :
    print (f"Suhu badan anda tidak normal")

elif SuhuBadan >= 36 and SuhuBadan <= 37 :
    print (f"Suhu badan anda normal")

else: 
    print(f"Anda mengalami demam")

engine = pyttsx3.init()

#rate 
rate = engine.getProperty("rate")
engine.setProperty("rate",150)
#print(rate)

#volume 
volume = engine.getProperty("volume")
engine.setProperty("volume",1)
#print("volume is {0}".format(volume))

#male female voices
voices = engine.getProperty("voices")
engine.setProperty("voice", voices[1].id) # for windows

def talk(text):
    engine.say(text)
    engine.runAndWait()

talk(f"good morning {name}")
talk(f"this is the data of your medical check up")
talk(f"your body temperature {temperature} celcius")
talk(f"your heartrate {heartrate}")
talk(f"your blood pressure {systole} and {diastole}")

